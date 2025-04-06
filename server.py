from flask import Flask, render_template, request, send_file
from gtts import gTTS
from googletrans import Translator
import os
from openpyxl import Workbook
from io import BytesIO
from openpyxl import load_workbook

app = Flask(__name__)

UPLOAD_FOLDER = 'static'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

EXCEL_FILE_PATH =  r'D:\lilproject\audio_gen\generation.xlsx'   # Update this to the actual path

def get_generated_text(input_text):
    wb = load_workbook(EXCEL_FILE_PATH)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == input_text:
            return row[1]
    return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_text():
    if request.method == 'POST':
        input_text = request.form['text']
        target_language = request.form['language']

        generated_text = get_generated_text(input_text)

        if generated_text is None:
            return "No generated text found for the input", 404

        translator = Translator()
        if target_language != 'en':
            translated_text = translator.translate(generated_text, src='en', dest=target_language).text
        else:
            translated_text = generated_text

        tts = gTTS(text=translated_text, lang=target_language, slow=False)
        audio_filename = f"output_{target_language}.mp3"
        audio_path = os.path.join(app.config['UPLOAD_FOLDER'], audio_filename)
        tts.save(audio_path)

        excel_filename = f"conversion_details_{target_language}.xlsx"
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Conversion Data"
        ws.append(["Input Text", "Generated Text", "Language", "Audio File"])
        ws.append([input_text, generated_text, target_language, audio_filename])
        wb.save(excel_path)

        audio_url = f"/static/{audio_filename}"
        excel_url = f"/static/{excel_filename}"

        return render_template('index.html', audio_url=audio_url, excel_url=excel_url)

if __name__ == '__main__':
    app.run(debug=True)
